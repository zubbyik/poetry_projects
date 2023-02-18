from playwright.sync_api import Page, expect
from pathlib import Path
import openpyxl
import re


data = set()
company_numbers = []
dataframe = openpyxl.load_workbook(Path().joinpath('land_use_reg.xlsx'))
dataframe1 = dataframe.active
# cell = dataframe1.cell(row=2, column=1)
# print('hello')
# print("result: \n"+cell.value)
for row in range(0, dataframe1.max_row):
    for col in dataframe1.iter_cols(1, dataframe1.max_column):
        data.add(col[row].value)

# reg = ["ELDRITCH LIMITED",
# "QUEENTON LIMITED",
# "DUNBAR VENTURES CORP.",
# "GESTRUST SA",
# "BARMOUTH LIMITED",
# "DREEMGARROW LIMITED",
# "KAA CONSULTING S.A. BVI"]

url = 'https://find-and-update.company-information.service.gov.uk/'
def test_land_reg_two(page: Page):
    page.goto(url)
    page.get_by_text("Reject analytics cookies").click()
    search_box = page.get_by_placeholder("Start here...")
    for item in data:
        search_box.fill(item)
        page.get_by_role("button", name="Search").click()
        try:
            search_link = page.get_by_role("link", name=item[:8]).first
            result = page.get_by_text(item[:8]).first
            not_found = page.get_by_text("No results found")
            if(search_link.is_visible() or search_link.text_content() != "" or not_found or result.is_visible()):
                search_link.click()
                comp_number = page.locator("#company-number").get_by_text(re.compile("^OE\d{6}")).text_content()
                if(comp_number != ''):
                    company_numbers.append(comp_number)
                page.goto(url)
            else:
                page.goto(url)
        except Exception as e:
            print("Link not found due to " + e.__str__())
            page.goto(url)
    with open("company_result.txt", 'w') as company_result:
        for item in comp_number:
            company_result.write(item)
        company_result.close()
    page.close()