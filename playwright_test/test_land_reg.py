from playwright.sync_api import Page, expect
import  openpyxl
from pathlib import Path




def test_land_use_act_and_scrape_the_text(page: Page):
    data = []
    dataframe = openpyxl.load_workbook(Path().joinpath('land_use_reg.xlsx'))
    dataframe1 = dataframe.active
    # cell = dataframe1.cell(row=2, column=1)
    # print('hello')
    # print("result: \n"+cell.value)
    for row in range(0, dataframe1.max_row):
        for col in dataframe1.iter_cols(1, dataframe1.max_column):
            data.append(col[row].value)

    # Get to the search Page
    # page.goto("https://find-and-update.company-information.service.gov.uk/")
    # page.get_by_role("button", name="Reject analytics cookies").click()
    # page.get_by_role("button", name="Hide this message").click()

    # print("\n"+heading)
    # page.close()
    for item in data:
        page.get_by_placeholder("Start here...").click()
        page.get_by_placeholder("Start here...").fill(item)
        page.get_by_role("button", name="Search").click()
        heading = page.get_by_text("OE005420 - Registered on 28 November 2022").text_content()